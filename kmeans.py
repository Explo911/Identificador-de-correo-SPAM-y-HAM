import os

import openpyxl
import preprocesador
import math


def evaluate(k, distances):
    dist_ordered = sorted(distances, key=lambda x: x[1])
    # print(dist_ordered)
    eval = dist_ordered[0:k]
    count = 0
    for p, v in eval:
        if p == "H":
            count += 1
    if count > k / 2:
        # print("Predicted value H")
        return "H"
    # print("Predicted value S")
    return "S"


def find_distance(P, Q):
    distance = 0
    for i in range(len(Q)):
        distance += (P[i] - Q[i]) ** 2
    distance = math.sqrt(distance)
    return distance


def confusion(real, predict):
    TP = 0
    FN = 0
    FP = 0
    TN = 0
    for i in range(len(real)):
        if real[i] == "H":
            if predict[i] == "H":
                TP += 1
            elif predict[i] == "S":
                FN += 1
        elif real[i] == "S":
            if predict[i] == "S":
                TN += 1
            elif predict[i] == "H":
                FP += 1
    print(" | H | S ")
    print("---------")
    print(f"H| {TP} | {FN}")
    print("---------")
    print(f"S| {FP} | {TN}")
    precision = TP / (TP + FP)
    sensibilidad = TP / (TP + FN)
    exactitud = (TP + TN) / (TP + TN + FN + FP)
    especificidad = TN / (TN + FP)
    tasaError = (FP + FN) / (TP + TN + FP + FN)
    presicionF = 2 * ((precision * sensibilidad) / (precision + sensibilidad))
    print(
        f"Precision: {precision}\nSensibilidad: {sensibilidad}\nExactitud: {exactitud}\nEspecificidad: {especificidad}\nTasa de Error: {tasaError}\nPresicion F: {presicionF}")
    return precision


def main(TFIDF_path, test_folder, test_results):
    preprocesador.main(test_folder, test_results, normalizar=True)
    tfidf = openpyxl.load_workbook(f"{TFIDF_path}/MatrizTF-IDF.xlsx")
    tf = tfidf.active
    test_tfidf = openpyxl.load_workbook(f"{test_results}/MatrizTF-IDFNormStems.xlsx")
    test_tf = test_tfidf.active
    rows = list(tf.rows)
    stems = [cell.value for cell in rows[0]]
    stems = stems[1:]
    # print(stems)
    trained_data = []
    for row in rows[1:]:
        vec = [cell.value for cell in row]
        vec = vec[1:]
        key = vec[-1]
        vec = vec[:-1]
        trained_data.append((key, vec))
    test_rows = list(test_tf.rows)
    test_stems = [cell.value for cell in test_rows[0]]
    test_stems = test_stems[1:]
    # print(test_stems)
    vectors = []
    for row in test_rows[1:]:
        vec = [cell.value for cell in row]
        vec = vec[1:]
        vectors.append(vec)
    # print(vectors)
    test_vectors = []
    for vec in vectors:
        temp = []
        for stem in stems:
            if stem in test_stems:
                pos = test_stems.index(stem)
                temp.append(vec[pos])
            else:
                temp.append(0)
        test_vectors.append(temp)
    # print(test_vectors)
    prediction3 = []
    prediction5 = []
    for vec in test_vectors:
        temp = []
        for k, v in trained_data:
            dist = find_distance(vec, v)
            temp.append((k, dist))
        prediction3.append(evaluate(3, temp))
        prediction5.append(evaluate(5, temp))
    real_values = []
    test_files = os.listdir(test_folder)
    for file in test_files:
        if "HAM" in file:
            real_values.append("H")
        elif "SPAM" in file:
            real_values.append("S")
    print("----- K = 3 ------")
    p1 = confusion(real_values, prediction3)
    print("----- K = 5 ------")
    p2 = confusion(real_values, prediction5)
    if p1 > p2:
        print("Se recomienda usar k = 3")
    else:
        print("Se recomienda usar k = 5")


def choose(correo, path_tfidf):
    correo_procesado = preprocesador.minipreprocesador(correo)
    matrixtfidf = openpyxl.load_workbook(f"{path_tfidf}/MatrizTF-IDF.xlsx")
    tf = matrixtfidf.active
    vector = []
    rows = list(tf.rows)
    stems = [cell.value for cell in rows[0]]
    stems = stems[1:]
    for stem in stems:
        if stem in correo_procesado:
            vector.append(correo_procesado[stem])
        else:
            vector.append(0)
    print(vector)
    trained_data = []
    for row in rows[1:]:
        vec = [cell.value for cell in row]
        vec = vec[1:]
        key = vec[-1]
        vec = vec[:-1]
        trained_data.append((key, vec))

    temp = []
    for k, v in trained_data:
        dist = find_distance(vector, v)
        temp.append((k, dist))
    prediction = evaluate(5, temp)
    return prediction
